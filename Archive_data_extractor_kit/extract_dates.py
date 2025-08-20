import pandas as pd
import numpy as np
import re

# === USER STEP: Specify input and output Excel file names ===
input_file = "data.xlsx"
output_file = "processed_file.xlsx"
    
# === STEP 1: Load Excel file and preserve leading zeros ===
# Preview first row to identify column names
preview = pd.read_excel(input_file, nrows=0)
first_col_name = preview.columns[0]

# Load the entire Excel file, treating the first two columns as strings to preserve leading zeros (e.g., folder numbers like 001)
df = pd.read_excel(input_file, dtype={0: str, 1: str})

# Drop fully empty rows
# Resets the index after cleaning
df.dropna(how='all', inplace=True)
df.reset_index(drop=True, inplace=True)

#----------------------------------------------------------------------------------------------------------
# === STEP 2: Mark section titles based on heuristic rules ===
# This section detects rows that act as section headers (titles), not actual data

for i in range(1, len(df) - 1):  # avoid index error
    curr_a = str(df.iloc[i, 0]).strip() if not pd.isna(df.iloc[i, 0]) else ""
    curr_b = str(df.iloc[i, 1]).strip() if not pd.isna(df.iloc[i, 1]) else ""
    curr_c = str(df.iloc[i, 2]).strip() if not pd.isna(df.iloc[i, 2]) else ""

    prev_b = str(df.iloc[i - 1, 1]).strip() if not pd.isna(df.iloc[i - 1, 1]) else ""
    next_b = str(df.iloc[i + 1, 1]).strip() if not pd.isna(df.iloc[i + 1, 1]) else ""
    next_a = str(df.iloc[i + 1, 0]).strip() if not pd.isna(df.iloc[i + 1, 0]) else ""

    if (
        curr_a != "" and
        curr_b == "" and
        curr_c != "" and
        (
            prev_b != "" or prev_b == ""  # optional: check that prev isn't another section
        ) and
        (re.fullmatch(r"\d{6}", next_a) or next_b != "")  # location or folder after
    ):
        df.at[i, df.columns[1]] = "##"  # mark as section title

#----------------------------------------------------------------------------------------------------------
# === STEP 3: Identify bolded subtitle rows using OpenPyXL ===
from openpyxl import load_workbook

# Load the original Excel file to preserve bold formatting info
wb_input = load_workbook(input_file)
ws_input = wb_input.active

# Collect bolded subtitle row indices
bold_subtitle_rows = set()

# Loop through each row in Excel and identify bolded subtitles
for idx, row in enumerate(ws_input.iter_rows(min_row=2), start=1):
    folder_cell = row[1].value
    desc_cell = row[2]
    is_bold = desc_cell.font.bold if desc_cell.font else False

    if (not folder_cell or str(folder_cell).strip() == "") and is_bold:
        bold_subtitle_rows.add(idx)

#-------------------------------------------------------------------------------------------------

# === STEP 4: Merge continuation rows into previous rows ===
rows_to_drop = []

for i in range(1, len(df)):
    val_b = df.iloc[i, 1]
    val_c = df.iloc[i, 2]

    raw_b = str(val_b).strip() if not pd.isna(val_b) else ""
    raw_c = str(val_c).strip() if not pd.isna(val_c) else ""

# If this is a continuation row (no folder # but has content)
    if (
        raw_b == "" and
        raw_c != "" and
        (i + 1) not in bold_subtitle_rows  # Excel rows are 1-indexed
    ):
        prev_val = str(df.iloc[i - 1, 2]).rstrip()
        merged = prev_val + " " + raw_c
        df.at[i - 1, df.columns[2]] = " ".join(merged.split())
        rows_to_drop.append(i)

# Drop merged rows and reindex
df.drop(index=rows_to_drop, inplace=True)
df.reset_index(drop=True, inplace=True)

#----------------------------------------------------------------------------------------------------------

# === STEP 5: Compile regex for comprehensive date extraction ===
date_pattern = re.compile(
    r"(?i)("
    
    # 0. Month DD, YYYY – Month DD, YYYY
    r"\b(?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|"
    r"Jul(?:y)?|Aug(?:ust)?|Sep(?:t(?:ember)?)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)"
    r"\s+\d{1,2},\s+\d{4}\s*[-–—/]\s*"
    r"(?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|"
    r"Jul(?:y)?|Aug(?:ust)?|Sep(?:t(?:ember)?)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)"
    r"\s+\d{1,2},\s+\d{4}"
    r"|"

    # 0a. Month – (Month)? DD?,? YYYY  <-- GENERALIZED for formats like:
#     - June 23 – July 13, 1985
#     - June 23–30, 1985
#     - January – June 1995
#     - July – August 2004
#     - June 1978 - July 1979
#     - June 1978-79
#     - June - May 1979-80
    r"(?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|"
    r"Jul(?:y)?|Aug(?:ust)?|Sep(?:t(?:ember)?)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)"
    r"(?:\s+\d{1,2})?"
    r"(?:,\s*)?"
    r"(?:\s+\d{4})?"
    r"\s*[-–—/]\s*"
    r"(?:"
        r"(?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|"
        r"Jul(?:y)?|Aug(?:ust)?|Sep(?:t(?:ember)?)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)"
        r"(?:\s+\d{1,2})?"
        r"(?:,\s*)?"
        r"\s+(?:\d{2}|\d{4})(?:[-–—/]?(?:\d{2}))?"
        r"(?:\s*[-–—/]\s*\d{2,4})?"
    r"|"
        r"\d{2,4}"
    r")"
    r"|"

    # 0b. Month/Month YYYY (e.g., November/December 1987)
    r"(?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|"
    r"Jul(?:y)?|Aug(?:ust)?|Sep(?:t(?:ember)?)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)"
    r"\s*/\s*"
    r"(?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|"
    r"Jul(?:y)?|Aug(?:ust)?|Sep(?:t(?:ember)?)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)"
    r"\s+\d{4}"
    r"|"

    # 1. Season ranges (Fall 1975 – Spring 1982)
    r"(Spring|Summer|Fall|Winter)\s+\d{4}\s*[-–—/]\s*(Spring|Summer|Fall|Winter)\s+\d{4}"
    r"|"

    # 2. Fiscal or Academic Year YYYY or with ranges
    r"\b(?:Fiscal|Academic)\s+Year\s+\d{4}(?:[-/]\d{2,4})?"
    r"|"

    # 3. Year ranges: 1978–79 or 1978–1980 (or enclosed in parentheses)
    r"\(?\d{4}\s*[-–—/]\s*\d{2,4}\)?"
    r"|"

    # 4. FY or AY Ranges and Abbreviations (e.g., FY 1993–94, FY 92)
    r"\b(?:F\.?Y\.?|A\.?Y\.?)\.?\s*-?\s*\d{2,4}(?:[-/]\d{2,4})?\b"
    r"|"

    # 5. Full MM/DD/YY – MM/DD/YY
    r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\s*[-–—to]+\s*\d{1,2}[/-]\d{1,2}[/-]\d{2,4}"
    r"|"

    # 6. Month Day, Year (also handles hyphenated day ranges like June 5–7, 1984)
    r"\b(?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|"
    r"Jul(?:y)?|Aug(?:ust)?|Sep(?:t(?:ember)?)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)"
    r"(?:\s+\d{1,2}(?:-\d{1,2})?)?(?:,\s*|\s+)\d{4}"
    r"|"
    
    # 7. Month, Year variations (e.g., (10/1978, 10/10/1978, 10/1978-11/1979))
    r"\d{1,2}[/-]\d{2,4}\s*(?:[-–—/]|to)\s*\d{1,2}[/-]\d{2,4}"
    r"|"
    r"\d{1,2}[/-]\d{4}"
    r"|"
    r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}"
    r"|"

    # 8. MM/YY (e.g., 03/92)
    r"\d{1,2}[/-]\d{2}"
    r"|"

    # 9. Comma-separated years (e.g., (1991, 1992, 1993))
    r"\(?\d{4}(?:,\s*\d{4}(?:[-/]\d{2,4})?)*\)?"
    r"|"

    # 10. Standalone 4-digit year
    r"\b\d{4}\b"
    r"|"

    # 11. DD MMM YY or DD-MMM-YY (e.g., 12 Sep 96 or 11-July 98)
    r"\d{1,2}[-\s](?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|"
    r"Jul(?:y)?|Aug(?:ust)?|Sep(?:t(?:ember)?)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)[-\s]\d{2}"
    r"|"

    # 12. Month YY (e.g., July 93)
    r"(?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|"
    r"Jul(?:y)?|Aug(?:ust)?|Sep(?:t(?:ember)?)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)\s+\d{2}"

    r")"
)
#---------------------------------------------------------------------------------------------

# === STEP 6: Extract scope-related metadata like [4 folders], 3 copies, etc. ===
scope_pattern = re.compile(
    r"[\[\(]\s*\d+\s+folders?\s*[\]\)]"         # e.g., [3 folders] or (4 folders)
    r"|\b\d+\s+copies\b",                       # e.g., 3 copies, 2 copies
    re.IGNORECASE
)

df["Scope"] = ""

# Apply scope extraction
for i in range(len(df)):
    val = str(df.iloc[i, 2])
    match = scope_pattern.search(val)
    if match:
        scope_text = match.group().strip()
        df.at[i, "Scope"] = scope_text
        df.at[i, df.columns[2]] = val.replace(scope_text, "").strip()

#------------------------------------------------------------------------------------------------------
# === STEP 7: Function to extract valid date strings from descriptions ===
def extract_text(cell):
    """
    Extracts a valid date string from a given cell value using regex patterns,
    while applying multiple safety checks to avoid false positives.

    Returns:
        tuple: (cleaned_text_without_date, extracted_date)
    """
    if pd.isna(cell):
        return "", ""

    cell_str = str(cell).strip()
    match = date_pattern.search(cell_str)

    while match:
        matched = match.group().strip("()")

        # ─── FILTER X: Reject standalone 4-digit numbers outside year range ─────────────────────────────
        if re.fullmatch(r"\d{4}", matched):
            year_candidate = int(matched)
            if not (1800 <= year_candidate <= 2099):
                match = date_pattern.search(cell_str, match.end())
                continue
                
        # ─── FILTER 1: Reject known false-positive short ranges like "1-50", "12–202" ─────────────────────
        if re.fullmatch(r"\d{1,3}\s*[-–—/]\s*\d{1,3}", matched):
            match = date_pattern.search(cell_str, match.end())
            continue

        # ─── FILTER 2: Reject codes like "K-2-10488" where "2-1048" falsely matches ──────────────────────
        if re.fullmatch(r"\d{1,2}[-–—/]\d{4}", matched):
            if match.start() >= 2 and cell_str[match.start() - 2 : match.start()].upper() in ["K-", "D-", "R-"]:
                match = date_pattern.search(cell_str, match.end())
                continue

        # ─── FILTER 3: Ignore matches that are part of long numbers (e.g., 89673772) ─────────────────────
        if re.fullmatch(r"\d{7,}", matched):
            match = date_pattern.search(cell_str, match.end())
            continue

        # ─── FILTER 4: Skip ranges embedded within larger numeric values like "312–320" ──────────────────
        before = cell_str[max(0, match.start() - 1):match.start()]
        after = cell_str[match.end():match.end() + 1]

        if before.isdigit() and after.isdigit():
            match = date_pattern.search(cell_str, match.end())
            continue

        # ─── VALIDATION: Expand partial year ranges like "1978–79" into "1978–1979" ──────────────────────
        if re.fullmatch(r"(18|19|20)\d{2}\s*[-–—/]\s*(\d{2}|\d{4})", matched):
            parts = re.split(r"[-–—/]", matched)
            try:
                p1 = int(parts[0].strip())
                p2_raw = parts[1].strip()
                p2 = int(p2_raw) if len(p2_raw) == 4 else int(str(p1)[:2] + p2_raw)

                if 1800 <= p1 <= 2099 and 1800 <= p2 <= 2099:
                    matched = f"{p1}-{p2}"
                else:
                    match = date_pattern.search(cell_str, match.end())
                    continue
            except Exception:
                match = date_pattern.search(cell_str, match.end())
                continue

        # ─── VALIDATION: Single year — ensure it's within acceptable range ───────────────────────────────
        elif re.fullmatch(r"\d{4}\s*[-–—/]\s*\d{4}", matched):
            parts = re.split(r"[-–—/]", matched)
            try:
                y1 = int(parts[0].strip())
                y2 = int(parts[1].strip())

                if not (1800 <= y1 <= 2099 and 1800 <= y2 <= 2099):
                    match = date_pattern.search(cell_str, match.end())
                    continue
            except Exception:
                match = date_pattern.search(cell_str, match.end())
                continue

        # ─── CLEANING AND RETURN ────────────────────────────────────────────────────────────────────────
        before_text = cell_str[:match.start()].strip()
        after_text = cell_str[match.end():].strip()
        cleaned_text = (before_text + " " + after_text).strip()

        if not cleaned_text:
            return cell_str, matched  # Only date, no surrounding content

        return cleaned_text, matched

    return cell_str, ""  # No match found
#----------------------------------------------------------------------------------------------------------

# Apply date extraction to Column C
# Store result in new columns, then overwrite
df["Before_Date"], df["Date_And_After"] = zip(*df.iloc[:, 2].map(extract_text))
df.iloc[:, 2] = df["Before_Date"]
df.drop(columns=["Before_Date"], inplace=True)
                
#----------------------------------------------------------------------------------------------------------
# === STEP 8: Assign location number to each box (handle 35a, 35b, etc.) ===
df["LocationNumber"] = ""

i = 0
while i < len(df):
    box_val = df.iloc[i, 0]
    
    if pd.notna(box_val):
        box_str = str(box_val).strip()

        # Match box numbers like 35, 35a, 35b — but exclude 6-digit location numbers
        if re.fullmatch(r"\d+[a-zA-Z]?", box_str):
            location_row = None
            location_value = None
            discard_rows = []

            # Step 1: Scan next 2 rows to find location or discard
            for offset in [1, 2]:
                r = i + offset
                if r >= len(df):
                    continue
                raw_val = df.iloc[r, 0]
                if pd.isna(raw_val):
                    continue
                val = str(raw_val).strip()
                padded = val.zfill(6) if val.isdigit() else val

                if re.fullmatch(r"\d{6}", padded) and location_row is None:
                    location_row = r
                    location_value = padded
                else:
                    discard_rows.append(r)

            # Step 2: Assign and clear only if location was found
            if location_value:
                df.at[i, "LocationNumber"] = location_value
                df.at[location_row, df.columns[0]] = np.nan

            for r in discard_rows:
                df.at[r, df.columns[0]] = np.nan

            # Step 3: Only fill down location if it was found
            j = i + 1
            while j < len(df):
                next_val = df.iloc[j, 0]
                if pd.notna(next_val):
                    next_box_str = str(next_val).strip()
                    if re.fullmatch(r"\d+[a-zA-Z]?", next_box_str):
                        break  # new box number reached
                if location_value:
                    df.at[j, "LocationNumber"] = location_value
                j += 1

            i = j
        else:
            i += 1
    else:
        i += 1
#------------------------------------------------------------------------------------------
# === STEP 9: Fill down missing Box/Location values ===
box_col = df.columns[0]
loc_col = "LocationNumber"

# Loop through all rows
for i in range(1, len(df)):
    prev_box = str(df.at[i - 1, box_col]).strip() if pd.notna(df.at[i - 1, box_col]) else ""
    curr_box = str(df.at[i, box_col]).strip() if pd.notna(df.at[i, box_col]) else ""
    
    # Fill down Box Number if current is blank
    if curr_box == "":
        df.at[i, box_col] = prev_box

    # Fill down Location Number if missing and box hasn’t changed
    if (
        (pd.isna(df.at[i, loc_col]) or str(df.at[i, loc_col]).strip() == "")
        and (str(df.at[i, box_col]).strip() == prev_box)
    ):
        df.at[i, loc_col] = df.at[i - 1, loc_col]

#-----------------------------------------------------------------------------------------

# === STEP 10: Save final cleaned data to Excel ===
df.to_excel(output_file, index=False)
print(f"Done! Output saved to: {output_file}")

#-----------------------------------------------------------------------------------------
# === STEP 11: Apply bold formatting back to subtitles and section titles ===
from openpyxl import load_workbook
from openpyxl.styles import Font
import re

# Load the processed Excel file
wb = load_workbook(output_file)
ws = wb.active

# Loop through all rows (starting from row 2 to skip header)
for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
    folder_val = str(row[1].value).strip() if row[1].value else ""
    content_val = str(row[2].value).strip() if row[2].value else ""

    # Case 1: Section title (marked with ## in column B)
    if re.match(r"#+", folder_val):
        cleaned = re.sub(r"^#+\s*", "", folder_val)
        row[1].value = cleaned
        row[2].font = Font(bold=True)

    # Case 2: Subtitle row (bolded, no folder number)
    elif folder_val == "" and content_val != "":
        row[2].font = Font(bold=True)

# Save the updated Excel file
wb.save(output_file)
